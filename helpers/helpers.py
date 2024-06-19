import secrets
from datetime import datetime
from tempfile import NamedTemporaryFile
from django.utils.encoding import smart_str
from django.http import HttpResponse
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

# Función para agregar datos y calcular totales
def agregar_datos_y_totales(ws, df, start_row):
    meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    total_por_mes = [0] * 12
    
    # Agrupar datos por concepto
    df_agrupado = df.groupby(['concept', 'month']).sum().reset_index()

    # Crear un diccionario para almacenar los valores agrupados
    data_dict = {}
    for _, row in df_agrupado.iterrows():
        if row['concept'] not in data_dict:
            data_dict[row['concept']] = [None] * 12
        data_dict[row['concept']][row['month'] - 1] = row['totalQuantity']

    # Agregar datos a la hoja de cálculo
    for concept, values in data_dict.items():
        total_por_mes = [total_por_mes[i] + (values[i] if values[i] is not None else 0) for i in range(12)]
        total_fila = sum(value for value in values if value is not None)
        ws.append([concept] + values + [total_fila])

    # Calcular el total por columna
    total_fila_final = ['Total'] + total_por_mes + [sum(total_por_mes)]
    ws.append(total_fila_final)

    # Aplicar formato de moneda a los valores
    for row in ws.iter_rows(min_row=start_row, max_row=start_row + len(data_dict) + 1, min_col=2, max_col=14):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '"$"#,##0.00'

def archivo_mensual(data, header_translation):

    # Crear un borde delgado
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"
    ws.append([None, None])

    # Add the headers to the worksheet
    headers = list(data[0].keys())
    translated_headers = [header_translation[header] for header in headers]
    ws.append([None]+translated_headers)

    # Add the data to the worksheet
    for entry in data:
        row = []
        for header in headers:
            try:
                cell_value = entry[header]
            except KeyError:
                # If the key doesn't exist, fill with empty value
                cell_value = ""
            row.append(cell_value)
        ws.append([None]+row)

    # Apply formatting to the 'Monto' column (assumed to be the 6th column)
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=8, max_row=len(data) + 2):
        for cell in row:
            cell.number_format = '"$"#,##0.00'

    # Apply formatting to the 'Monto' column (assumed to be the 6th column)
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=len(data[0])+1, max_row=len(data) + 2):
        for cell in row:
            cell.border = thin_border

    # Ajustar automÃ¡ticamente el ancho de las columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Obtener la letra de la columna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 1)
        ws.column_dimensions[column].width = adjusted_width

    return wb

def archivo_anual(ingresos_por_mes, gastos_por_mes, inventario_por_mes, gastos_generales_por_mes):
    # Crear un archivo de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Anual"

    # Agregar datos y totales de ingresos
    ws.append(["Ingresos"])
    ws.append(["Concepto", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre", "Total"])
    
    if len(ingresos_por_mes) > 0:
        ingresos_df = pd.DataFrame(ingresos_por_mes)
        agregar_datos_y_totales(ws, ingresos_df, start_row=3)

    # Agregar separador
    ws.append([])

    # Agregar datos y totales de egresos
    ws.append(["Egresos"])
    ws.append(["Concepto", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre", "Total"])

    if len(gastos_por_mes) > 0:
        egresos_df = pd.DataFrame(inventario_por_mes+gastos_por_mes)
        agregar_datos_y_totales(ws, egresos_df, start_row=ws.max_row + 1)

    # Agregar separador
    ws.append([])

    # Agregar datos y totales de gastos generales
    ws.append(["Gastos Generales"])
    ws.append(["Concepto", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre", "Total"])
    if len(gastos_generales_por_mes) > 0:
        gastos_generales_df = pd.DataFrame(gastos_generales_por_mes)
        agregar_datos_y_totales(ws, gastos_generales_df, start_row=ws.max_row + 1)

    wb.close()
    return wb

def resumen_evento(event, gastos, ingresos, inventario):
    # Crear un archivo de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Evento"

    # Estilos
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    center_aligned_text = Alignment(horizontal="center")
    currency_format = '"$"#,##0.00'

    # Información del evento en una sola fila
    headers = ["Cliente",	"Tipo de evento",	"Ubicación",	"No. Asistentes",	"Precio evento", "ID Evento",	"Estado", "Fecha"]
    values = [event['name'], event['type'], event['location'], event['num_of_people'], event['cost'], event['id_event'], event['state'], event['date']]

    headers2 =[" ",	"Total ingresos",	"Total egresos",	"Renta de salón", "Utilidad","Margen",	"Margen sin Salon",]
    values2 = [" ", event['in'], event['out'], event['salon'], event['in']-event['out'], (event['in']-event['out'])/event['out'], (event['in']-event['out'])/(event['out']-event['salon']) ]

    ws.append(headers)
    ws.append(values)
    ws.append([None, None])
    ws.append(headers2)
    ws.append(values2)



    # Aplicar formato de moneda a las celdas E2, J2, K2, L2, M2, N2
    ws['E2'].number_format = currency_format
    ws['B5'].number_format = currency_format
    ws['C5'].number_format = currency_format
    ws['D5'].number_format = currency_format
    ws['E5'].number_format = currency_format

    # Aplicar formato de porcentaje a las celdas N2 y O2
    ws['F5'].number_format = '0.00%'
    ws['G5'].number_format = '0.00%'


    # Agregar descripción
    ws.append(["Descripcion:"])
    ws.append([""])

    # Ingresos
    ws.append(["Ingresos"])
    ws.append(["Nombre del Cliente", "Fecha del Evento", "Importe", "IVA", "Total", "Quien Realizo el Pago", "Concepto", "Fecha del Pago", "Folio"])

    for ingreso in ingresos:
        total = ingreso['quantity']
        iva = 0  # Suponiendo que el IVA es 0 en este ejemplo
        ws.append([ingreso['payer'], event['date'], ingreso['quantity'], iva, total, ingreso['payer'], ingreso['concept'], ingreso['date'], ingreso['id_ticket']])

    # Aplicar estilo a los encabezados de ingresos
    for cell in ws[ws.max_row - len(ingresos) - 1]:
        cell.font = header_font
        cell.fill = PatternFill(start_color="00DC32", end_color="00DC32", fill_type="solid")
        cell.alignment = center_aligned_text

    # Egresos
    ws.append(["Egresos"])
    ws.append(["Nombre del Cliente", "Fecha del Evento", "Proveedor", "Concepto", "Monto", "IVA", "Total", "Fecha del Pago", "Folio"])

    for gasto in gastos:
        total = gasto['amount']
        iva = 0  # Suponiendo que el IVA es 0 en este ejemplo
        ws.append([event['name'], event['date'], gasto['buyer'], gasto['concept'], gasto['amount'], iva, total, gasto['date'], gasto['invoice']])

    # Aplicar estilo a los encabezados de egresos
    for cell in ws[ws.max_row - len(gastos) - 1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_aligned_text

    # Aplicar formato de moneda a los valores
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, min_col=3, max_col=7):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = currency_format


    # Inventario
    ws.append(["Inventario"])
    ws.append(["Parte", "Proveedor", "Concepto", "Cantidad", "Importe", "Fecha", "Factura"])

    for item in inventario:
        ws.append([item['portion'], item['buyer'], item['concept'], item['quantity'], item['amount'], item['date'], item['invoice']])

    # Aplicar estilo a los encabezados de inventario
    for cell in ws[ws.max_row - len(inventario) - 1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_aligned_text

    # Aplicar formato de moneda a los valores del inventario
    for row in ws.iter_rows(min_row=ws.max_row - len(inventario) + 1, max_row=ws.max_row, min_col=5, max_col=5):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = currency_format


    #Ajustar el ancho de las columnas automáticamente
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Centrar el texto en todas las celdas
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center_aligned_text

    # Guardar el archivo
    return wb

def search(query, collection):
    return list(collection.find(query))

def searchWithProjection(query, projection, collection, errorMessage):
    res = list(collection.find (query, projection))
    if not res:
        status = 404
        res = [{'message' : errorMessage}]
        return (res, status)
    else:
        return (res, 200) 
    
def searchWithPagination(pipeline, projection, collection, errorMessage):
    res = collection.aggregate(pipeline, projection)
    if not res:
        status = 404
        res = [{'message' : errorMessage}]
        return (res, status)
    else:
        return (res, 200) 
    
def deleteExtraQueries(data, expected_keys):
        # Get the keys of the JSON object
    data_keys = data.keys()
    
    # Check if all keys are in the expected keys list
    for key in data_keys:
        if key not in expected_keys:
            del data[key]
    return data

def getIDEvento(data, bytes):
    #check if month and day less than 10
    month = '0' + str(data['month']) if data['month'] < 10 else str(data['month'])
    day = '0' + str(data['day']) if data['day'] < 10 else str(data['day'])
    date = str(data['year'])[2:] + month + day
    random_string = secrets.token_hex(bytes)  # Generate a random hex string of 16 bytes (32 characters)
    id_evento = random_string + str(date)
    
    return ''.join(id_evento)

# check if contains all the expected_keys and not others
def checkData(data, keys, types):
    for key in keys:
        if key not in data:
            return (False, "Falta llave {} en la informacion enviada".format(key))
        if isinstance(types[key], list):
            isGood = False
            for type_ in types[key]:
                if isinstance(data[key], type_):
                    isGood = True
            if not isGood:
                return (False, "Tipo de data incorrecto {} en llave {}".format(type(data[key]), type_))
        elif not isinstance(data[key], types[key]):
            print(type(data[key]),types[key] )
            return (False, "Tipo de data incorrecto {} en llave {}".format(type(data[key]), key))
    return (True, "Los datos se ven bien")


# check if contains keys between a range
def check_keys(data, expected_keys):
    # Get the keys of the JSON object
    data_keys = data.keys()
    
    # Check if all keys are in the expected keys list
    for key in data_keys:
        if key not in expected_keys:
            return False
    return True

def updateData(collection, query, updateQuery):
    res = collection.update_one(query, updateQuery)
    response = {}

    update_result = res.raw_result

    response['message'] = "Se modifico satisfactoriamente {} registros".format(update_result["nModified"])
    response['result'] = update_result["nModified"]

    return response

def checkForChangeOfID(data, collection):
    id_evento = data['id_evento']
    old_id = id_evento
    
    res = collection.find_one({'id_evento' : id_evento})
    if res['location'] != data['location'] or res['date'] != data['date']:
        return (getIDEvento(data), old_id)
    return (id_evento, id_evento)

def getDate():
    present_date = datetime.now().date()
    present_date_str = present_date.strftime("%d-%m-%Y")
    return present_date_str

def generateTicketNumber(date):
    random_start = secrets.token_hex(1)  
    randon_end = secrets.token_hex(1)
    new_date = date.split('-')
    new_date[2] = new_date[2][:2]
    print(new_date)
    id_ticket = [random_start] + new_date  + [randon_end]
    return ''.join(id_ticket)

def generateIDTicket(date):
    random_end = secrets.token_hex(2)  
    id_ticket = date + random_end
    return ''.join(id_ticket)

def findAbono(query, collection):
    res = list(collection.find(query))
    status = 200
    if not res:
        status = 404
        res = {'message' : 'Pagos no encontrados'}
    return (res, status)



def returnExcel(wb, filename) -> HttpResponse:

    # Read the entire file as bytes
    new_buffer = BytesIO()
    wb.save(new_buffer)
    new_buffer.seek(0)
    # Create an HttpResponse with the Excel file as attachment
    response = HttpResponse(new_buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'

    return response

